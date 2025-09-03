import streamlit as st
import pandas as pd
import os
import re
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

# Pillow / image support (graceful if missing)
try:
    from openpyxl.drawing.image import Image as XLImage
    PIL_OK = True
except ImportError:
    XLImage = None
    PIL_OK = False


# -------------------------- CORE UTILITIES (Adapted for Streamlit) -----------------------------

def clean_header(headers):
    """Cleans and normalizes DataFrame headers."""
    return [str(h).strip().replace("\n", " ") for h in headers]

def process_csv(uploaded_file, filename: str) -> pd.DataFrame:
    """Read CSV from Streamlit's UploadedFile object, normalize headers, and add Server column."""
    if not filename:
        raise ValueError("Cannot determine server name from an empty filename.")

    # Read headers first to validate
    headers_df = pd.read_csv(uploaded_file, nrows=0)
    uploaded_file.seek(0) # Reset file pointer after reading headers
    headers = clean_header(list(headers_df.columns))

    if len(headers) not in [19, 20]:
        raise ValueError(f"Error: Expected 19 or 20 headers, but found {len(headers)}.")

    if len(headers) == 20:
        headers = headers[:19] + ["Tag"]
    else: # len is 19
        headers = headers + ["Tag"]

    # Read the full data
    df = pd.read_csv(uploaded_file, header=None, skiprows=1, dtype=str)
    if len(df.columns) != 20:
        raise ValueError(f"Error: Expected 20 columns in data rows, but found {len(df.columns)}.")

    df.columns = headers

    # Add Server column from the original filename
    server_name = filename.split('_')[0].split(' ')[0]
    df['Server'] = server_name

    return df

# ---------- Numeric coercion helpers (Unchanged) ---
_NUM_RE = re.compile(r'^[+-]?\d+([.]\d+)?$')

def _maybe_to_number(val):
    if val is None: return None
    if isinstance(val, (int, float)): return val
    s = str(val).strip()
    if s == "": return ""
    s_clean = s.replace(",", "")
    if not _NUM_RE.match(s_clean): return val
    if "." not in s_clean and "e" not in s_clean.lower():
        if len(s_clean.lstrip("+-")) > 15: return s
        try: return int(s_clean)
        except (ValueError, TypeError): return val
    try: return float(s_clean)
    except (ValueError, TypeError): return val

def _set_num_format(cell, value):
    if isinstance(value, int): cell.number_format = "0"
    elif isinstance(value, float): cell.number_format = "0.00"

def _autofit(ws, scan_rows=200, min_w=10, max_w=60, skip_letters=None):
    skip = set(skip_letters or [])
    for col in ws.columns:
        try:
            letter = get_column_letter(col[0].column)
            if letter in skip: continue
            max_len = 0
            for cell in col[:scan_rows]:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max(min_w, max_len + 2), max_w)
        except Exception: pass

# -------------------------- EXCEL GENERATION (Adapted for Streamlit) --------------------------------

def export_orderbook_with_formulas(df: pd.DataFrame, user_ids: list[str], mtm_image_file: io.BytesIO | None = None) -> tuple[Workbook, dict]:
    """Create the HEDGES workbook object; returns (Workbook, stats_dict)."""
    if 'Status' in df.columns:
        df['Status'] = df['Status'].fillna('').str.upper()
        status_col = 'Status'
    elif 'Tag' in df.columns:
        df['Tag'] = df['Tag'].fillna('').str.upper()
        status_col = 'Tag'
    else:
        raise ValueError("No 'Status' or 'Tag' column found in CSV.")

    # Formula definitions (Unchanged)
    formula_columns = {
        "C": ["=RIGHT(B2,2)", "=RIGHT(B{r},2)"],
        "K": ["=IF(F2=\"SELL\",-H2,H2)", "=IF(F{r}=\"SELL\",-H{r},H{r})"],
        "M": ["=IF(C2=\"CE\",K2*1,K2*0)", "=IF(C{r}=\"CE\",K{r}*1,K{r}*0)"],
        "N": ["=IF(C2=\"PE\",K2*1,K2*0)", "=IF(C{r}=\"PE\",K{r}*1,K{r}*0)"],
        "O": ["=M2", "=O{prev}+M{r}"],
        "P": ["=N2", "=P{prev}+N{r}"],
        "S": ["=IF(M2<0,0,M2)", "=IF(M{r}<0,0,M{r})+S{prev}"],
        "T": ["=IF(M2<0,M2,0)", "=IF(M{r}<0,M{r},0)+T{prev}"],
        "U": ["=IF(P2<0,0,P2)", "=IF(N{r}<0,0,N{r})+U{prev}"],
        "V": ["=IF(N2<0,N2,0)", "=IF(N{r}<0,N{r},0)+V{prev}"],
        "W": ["=IFERROR(ABS(S2)/ABS(T2),0)", "=IFERROR(ABS(S{r})/ABS(T{r}),0)"],
        "X": ["=IFERROR(ABS(U2)/ABS(V2),0)", "=IFERROR(ABS(U{r})/ABS(V{r}),0)"]
    }
    formula_cols_letters = list(formula_columns.keys())

    # Custom headers (Unchanged)
    custom_headers = [
        "SNO", "Symbol", "CE/PE", "Order Time", "Order ID", "Transaction", "Order Type",
        "Quantity", "Price", "Exchange Time", "B/S", "Avg Price", "CE", "PE",
        "CECum", "PECum", "User Alias", "User ID",
        "CE Buy", "CE Sell", "PE Buy", "PE Sell", "CE B/S", "PE B/S"
    ]

    wb = Workbook(); wb.remove(wb.active)

    # Sheet 1: Full Orderbook
    ws_full = wb.create_sheet("Orderbook Full")
    for c_idx, header in enumerate(df.columns, start=1):
        ws_full.cell(row=1, column=c_idx).value = header
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            val = _maybe_to_number(value)
            cell = ws_full.cell(row=r_idx, column=c_idx)
            cell.value = val
            _set_num_format(cell, val)
    _autofit(ws_full) # Autofit the full sheet as well

    # Per-user sheets
    created_sheets = []
    for uid in [u.strip() for u in user_ids if u.strip()]:
        user_df = df[(df.get('User ID', pd.Series(dtype=str)) == uid) & (df[status_col] == 'COMPLETE')]
        ws = wb.create_sheet(uid)
        created_sheets.append(uid)

        for c_idx, header in enumerate(custom_headers, start=1):
            ws.cell(row=1, column=c_idx).value = header

        start_row = 2
        for r_off, row in enumerate(user_df.itertuples(index=False), start=0):
            r = start_row + r_off
            for c_idx, value in enumerate(row, start=1):
                val = _maybe_to_number(value)
                cell = ws.cell(row=r, column=c_idx)
                cell.value = val
                _set_num_format(cell, val)
            for col_letter, (f_first, f_next) in formula_columns.items():
                formula = f_first if r == 2 else f_next.replace('{r}', str(r)).replace('{prev}', str(r - 1))
                ws[f"{col_letter}{r}"] = formula

    # MTM sheet
    ws_mtm = wb.create_sheet("MTM", index=len(wb.worksheets))
    ws_mtm["A1"] = "MTM Image:"
    ws_mtm["A1"].font = ws_mtm["A1"].font.copy(bold=True)

    if mtm_image_file and PIL_OK:
        try:
            img = XLImage(mtm_image_file)
            target_width = 800
            if hasattr(img, 'width') and img.width:
                scale = target_width / float(img.width)
                img.width = int(img.width * scale)
                img.height = int(img.height * scale)
            ws_mtm.add_image(img, "A2")
        except Exception as e:
            ws_mtm["A3"] = f"(Could not load image: {e})"
    elif mtm_image_file and not PIL_OK:
        ws_mtm["A3"] = "(Pillow library not installed; cannot embed image.)"
    else:
        ws_mtm["A3"] = "(No image uploaded)"

    # Style user worksheets
    light_green = PatternFill(fill_type="solid", fgColor="C6EFCE")
    FORMULA_WIDTHS = {"C": 6, "K": 7, "M": 7, "N": 7, "O": 8, "P": 8, "S": 8, "T": 8, "U": 8, "V": 8, "W": 10, "X": 10}
    
    for sheet_name in created_sheets:
        ws = wb[sheet_name]
        last_row = ws.max_row
        for col_letter in formula_cols_letters:
            for r in range(2, last_row + 1):
                ws[f"{col_letter}{r}"].fill = light_green
        for col_letter, width in FORMULA_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width
        _autofit(ws, scan_rows=300, min_w=12, max_w=60, skip_letters=formula_cols_letters)

    stats = {
        "total_rows": len(df),
        "total_cols": len(df.columns),
        "sheets": created_sheets,
    }

    return wb, stats


# ------------------------------ STREAMLIT UI ---------------------------------
st.set_page_config(page_title="Hedges Generator", layout="centered")

st.title("📈 Hedges Generator")
st.markdown("Create Excel exports from your Orderbook CSV.")

# --- Initialize session state ---
if 'excel_data' not in st.session_state:
    st.session_state.excel_data = None
    st.session_state.stats = None
    st.session_state.download_name = None
if 'user_ids' not in st.session_state:
    st.session_state.user_ids = []


# --- UI for adding and managing User IDs ---
st.header("1. Enter User IDs")
st.markdown("Add User IDs one by one. They will be used to create individual sheets in the Excel file.")

# Input for new User ID
col1, col2 = st.columns([3, 1])
with col1:
    new_user_id = st.text_input("new_user_id", placeholder="Enter a User ID and click Add", label_visibility="collapsed")
with col2:
    if st.button("Add User ID", use_container_width=True):
        if new_user_id:
            user_to_add = new_user_id.strip().upper()
            if user_to_add not in st.session_state.user_ids:
                st.session_state.user_ids.append(user_to_add)
                st.toast(f"✅ Added '{user_to_add}'!", icon="🎉")
                st.rerun() # Rerun to clear input and update list
            else:
                st.toast(f"⚠️ '{user_to_add}' is already in the list.", icon="🚨")
        else:
            st.toast("Please enter a User ID.", icon="❗")


# Display current list of User IDs
if st.session_state.user_ids:
    st.write("Current User IDs:")
    with st.container(border=True):
        for i, user_id in enumerate(st.session_state.user_ids):
            col1, col2 = st.columns([4, 1])
            col1.code(user_id, language=None)
            if col2.button("❌", key=f"remove_{user_id}", help=f"Remove {user_id}", use_container_width=True):
                st.session_state.user_ids.pop(i)
                st.toast(f"🗑️ Removed '{user_id}'.")
                st.rerun()
else:
    st.info("No User IDs have been added yet.", icon="ℹ️")


# --- UI Form for uploads and generation ---
with st.form("hedges_form"):
    st.header("2. Upload Orderbook CSV")
    uploaded_csv = st.file_uploader(
        "Select a CSV file",
        type="csv",
        help="The orderbook export file from your system."
    )
    
    # Note: The User ID input has been moved outside the form for a better dynamic experience.
    # The form will read the list of IDs from the session state when submitted.

    st.header("3. Upload MTM Image (Optional)")
    if not PIL_OK:
        st.warning("Pillow library is not installed. Image uploads are disabled. `pip install pillow`")
        uploaded_image = None
    else:
        uploaded_image = st.file_uploader(
            "Select an image file for the MTM sheet",
            type=["png", "jpg", "jpeg", "bmp", "gif"]
        )
    
    st.markdown("---")
    submitted = st.form_submit_button("🚀 Generate HEDGES Excel")


# --- Processing Logic ---
if submitted:
    # Reset previous results
    st.session_state.excel_data = None
    st.session_state.stats = None
    st.session_state.download_name = None

    # Validation
    if not uploaded_csv:
        st.error("Please upload an Orderbook CSV file.")
    elif not st.session_state.user_ids: # MODIFIED: Check the list from session state
        st.error("Please add at least one User ID using the input above.")
    else:
        user_ids = st.session_state.user_ids # MODIFIED: Use the list from session state
        with st.spinner("Processing... Generating your Excel file..."):
            try:
                # 1. Process the CSV
                df = process_csv(uploaded_csv, uploaded_csv.name)
                
                # 2. Generate the Excel workbook in memory
                workbook, stats = export_orderbook_with_formulas(df, user_ids, uploaded_image)
                
                # 3. Save workbook to an in-memory buffer
                excel_buffer = io.BytesIO()
                workbook.save(excel_buffer)
                
                # 4. Store buffer and stats in session state for download
                st.session_state.excel_data = excel_buffer.getvalue()
                st.session_state.stats = stats
                
                base_name = os.path.splitext(uploaded_csv.name)[0]
                st.session_state.download_name = f"{base_name} HEDGES.xlsx"
                
            except Exception as e:
                st.error(f"An error occurred: {e}")
                st.exception(e) # Also print the full traceback for debugging

# --- Display Results and Download Button ---
if st.session_state.excel_data:
    st.success("✅ Generation Complete!")

    stats = st.session_state.stats
    st.markdown(f"Your file **{st.session_state.download_name}** is ready to download.")

    col1, col2, col3 = st.columns(3)
    col1.metric("Total Rows Processed", f"{stats['total_rows']:,}")
    col2.metric("Columns Detected", stats['total_cols'])
    col3.metric("User Sheets Created", len(stats['sheets']))
    
    st.download_button(
        label="⬇️ Download Excel File",
        data=st.session_state.excel_data,
        file_name=st.session_state.download_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )